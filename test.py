def dataprocess():
    #适用于300ms 采样率位100hz的数据

    import openpyxl
    import os
    #import csv
    ## turn txt file to csv file  
    #csvFile = open("data.csv",'w',newline='',encoding='utf-8')
    #writer = csv.writer(csvFile)
    #csvRow=[]

    #f = open(r'C:\Users\86136\Desktop\1msVt of temp.txt','r',encoding='GB2312')
    #for line in f:
    #    csvRow=line.split()
    #    writer.writerow(csvRow)
    #f.close()
    #csvFile.close()
    # data process
    os.chdir(r'C:\Users\86136\source\repos')
    wb = openpyxl.load_workbook("data2.xlsx")
    sheet = wb.active
    wb.create_sheet(index=1,title='res_value')
    ws = wb['res_value']
    # 1-4 号器件 放在 2-5列 50个数据
    for num in range (1,10):
        for row in range (126+num*500-1,131+num*500-1):#   +num*500
            for col in range (2,6):
                ws.cell(column=col, row=row-125-num*495+1, value=sheet.cell(column=col,row=row).value )  
    for row in range (126-1,131-1):#   +num*500
        for col in range (2,6):
            ws.cell(column=col, row=row-125+1, value=sheet.cell(column=col,row=row).value)
    # 5-8 号器件 放在 6-9列 50个数据
    for row in range (156-1,161-1):#   +num*500
        for col in range (6,10):
            ws.cell(column=col, row=row-155+1, value=sheet.cell(column=col-4,row=row).value)
    for num in range (1,10):
        for row in range (156+num*500-1,161+num*500-1):
            for col in range (6,10):
                ws.cell(column=col, row=row-num*495-155+1, value=sheet.cell(column=col-4,row=row).value )
    # 9-12 号器件 放在 10-13列 50个数据
    for row in range (186-1,191-1):
        for col in range (10,14):
            ws.cell(column=col, row=row-185+1, value=sheet.cell(column=col-8,row=row).value)
    for num in range (1,10):
        for row in range (186+num*500-1,191+num*500-1):
            for col in range (10,14):
                ws.cell(column=col, row=row-num*495-185+1, value=sheet.cell(column=col-8,row=row).value )
    # 13-16 号器件 放在 14-17列 50个数据
    for row in range (216-1,221-1):
        for col in range (14,18):
            ws.cell(column=col, row=row-215+1, value=sheet.cell(column=col-12,row=row).value)
    for num in range (1,10):
        for row in range (216+num*500-1,221+num*500-1):
            for col in range (14,18):
                ws.cell(column=col, row=row-num*495-215+1, value=sheet.cell(column=col-12,row=row).value )
    #计算电阻
    for row in range (1,51):
        for col in range (2,18):
            ws.cell(column=col,row=row,value=(5-float(ws.cell(column=col,row=row).value))/float(ws.cell(column=col,row=row).value))
    #51-52行创建两行0 
    for row in range (51,53):
        for col in range (2,18):
            ws.cell(column=col,row=row,value=0)
    #7-9周期求和 放在51行
    for row in range (31,46):
        for col in range (2,18):
            ws.cell(column=col,row=51,value=(ws.cell(column=col,row=51).value + float(ws.cell(column=col,row=row).value)))
    #求平均值
    for col in range (2,18):
       ws.cell(column=col,row=51,value=ws.cell(column=col,row=51).value/15)
    #求方差
    for row in range (31,46):
       for col in range (2,18):
           ws.cell(column=col,row=52,value=ws.cell(column=col,row=52).value + (float(ws.cell(column=col,row=row).value)-  ws.cell(column=col,row=51).value)**2)
    #wb.create_sheet(index=2,title='return_sheet')
    #ws1 = wb['return_sheet']  
    #for cle in range (7,10): #只能从1-3  后续改为7-9的
    #    for col in range (2,10):
    #        ws1.cell(column=col,row=cle,value=0)  #初始值为0 int类型  cle 的表格位置需要调整
    #        for plus in range (0,5): # 从250-290ms的数据   
    #            ws1.cell(column=col,row=cle,value=(ws1.cell(column=col,row=cle).value + float (ws.cell(column=col,row=25+plus+30*(cle-1)).value)))
    ## 数据求平均值


    wb.save("data2.xlsx")

    wb1 = openpyxl.load_workbook("data1.xlsx")
    sheet = wb1.active
    wb1.create_sheet(index=1,title='res_value')
    ws1 = wb1['res_value']
    # 1-4 号器件 放在 2-5列 50个数据
    for num in range (1,10):
        for row in range (126+num*500-1,131+num*500-1):#   +num*500
            for col in range (2,6):
                ws1.cell(column=col, row=row-125-num*495+1, value=sheet.cell(column=col,row=row).value )  
    for row in range (126-1,131-1):#   +num*500
        for col in range (2,6):
            ws1.cell(column=col, row=row-125+1, value=sheet.cell(column=col,row=row).value)
    # 5-8 号器件 放在 6-9列 50个数据
    for row in range (156-1,161-1):#   +num*500
        for col in range (6,10):
            ws1.cell(column=col, row=row-155+1, value=sheet.cell(column=col-4,row=row).value)
    for num in range (1,10):
        for row in range (156+num*500-1,161+num*500-1):
            for col in range (6,10):
                ws1.cell(column=col, row=row-num*495-155+1, value=sheet.cell(column=col-4,row=row).value )
    # 9-12 号器件 放在 10-13列 50个数据
    for row in range (186-1,191-1):
        for col in range (10,14):
            ws1.cell(column=col, row=row-185+1, value=sheet.cell(column=col-8,row=row).value)
    for num in range (1,10):
        for row in range (186+num*500-1,191+num*500-1):
            for col in range (10,14):
                ws1.cell(column=col, row=row-num*495-185+1, value=sheet.cell(column=col-8,row=row).value )
    # 13-16 号器件 放在 14-17列 50个数据
    for row in range (216-1,221-1):
        for col in range (14,18):
            ws1.cell(column=col, row=row-215+1, value=sheet.cell(column=col-12,row=row).value)
    for num in range (1,10):
        for row in range (216+num*500-1,221+num*500-1):
            for col in range (14,18):
                ws1.cell(column=col, row=row-num*495-215+1, value=sheet.cell(column=col-12,row=row).value )
    #计算电阻
    for row in range (1,51):
        for col in range (2,18):
            ws1.cell(column=col,row=row,value=(5-float(ws1.cell(column=col,row=row).value))/float(ws1.cell(column=col,row=row).value))
    #51-52行创建两行0 
    for row in range (51,53):
        for col in range (2,18):
            ws1.cell(column=col,row=row,value=0)
    #7-9周期求和 放在51行
    for row in range (31,46):
        for col in range (2,18):
            ws1.cell(column=col,row=51,value=(ws1.cell(column=col,row=51).value + float(ws1.cell(column=col,row=row).value)))
    #求平均值
    for col in range (2,18):
       ws1.cell(column=col,row=51,value=ws1.cell(column=col,row=51).value/15)
    #求方差
    for row in range (31,46):
       for col in range (2,18):
           ws1.cell(column=col,row=52,value=ws1.cell(column=col,row=52).value + (float(ws1.cell(column=col,row=row).value)-  ws1.cell(column=col,row=51).value)**2)
    wb1.save("data1.xlsx")