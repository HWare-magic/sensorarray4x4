from module1 import * 
from test import *
import openpyxl
import os
os.chdir(r'C:\Users\86136\source\repos')
#转到对应目录
import csv
import xlwt
import win32com.client as win32
txtexcel()
fname = r'C:\Users\86136\source\repos\data2.xls'
excel = win32.DispatchEx('Excel.Application')
#excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)
wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
wb.Close()                               #FileFormat = 56 is for .xls extension
excel.Application.Quit()

fname = r'C:\Users\86136\source\repos\data1.xls'
excel = win32.DispatchEx('Excel.Application')
#excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)
wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
wb.Close()                               #FileFormat = 56 is for .xls extension
excel.Application.Quit()

dataprocess()

wb = openpyxl.load_workbook("data2.xlsx")
wb1 = openpyxl.load_workbook("data1.xlsx")
ws= wb.get_sheet_by_name('res_value')
ws1=wb1.get_sheet_by_name('res_value')
wb.create_sheet(index=3,title="total")


wball=wb['total']
# ws1是 data1 为 50ppm rs  ws是data2 为200ppm rs
for col in range (2,18):
    for row in range (31,46):
        wball.cell(column=col-1,row=row-30,value=float(ws1.cell(column=col,row=row).value)/float(ws.cell(column=col,row=row).value))
for row in range (16,18):
    for col in range (1,17):
        wball.cell(column=col,row=row,value=0)
for row in range (1,16):
    for col in range (1,17):
        wball.cell(column=col,row=16,value=(wball.cell(column=col,row=16).value + float(wball.cell(column=col,row=row).value)))
for col in range (1,17):
    wball.cell(column=col,row=16,value=wball.cell(column=col,row=16).value/15)
for row in range (1,16):
       for col in range (1,17):
           wball.cell(column=col,row=17,value=wball.cell(column=col,row=17).value + (float(wball.cell(column=col,row=row).value)-  wball.cell(column=col,row=16).value)**2)
wb.save("data2.xlsx")
